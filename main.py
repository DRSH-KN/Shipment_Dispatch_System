from os import path, startfile, system
import time
from datetime import date, datetime
from math import isclose
import json 

from PyQt5 import QtWidgets, QtGui

import openpyxl
from jinja2 import Environment, FileSystemLoader
import pdfkit

from RFID import arduino
from WeighScale import weighScale
from ui import Ui_MainWindow


class MainApp(Ui_MainWindow):

    def gui_append(self):
        self.loginBtn.clicked.connect(self.loginBtnClick)
        self.movie= QtGui.QMovie('Resources/load.gif')
        self.movie.finished.connect(self.loadingDone)
        self.label_5.setMovie(self.movie)
        self.TloadCount = 0;
        self.CloadCount = 0;
        self.NloadIndex = 0;
        self.dashLogoutBtn.clicked.connect(self.dashLogoutBtnClick)
        self.dashBtn1.clicked.connect(self.dashBtn1Click)
        self.dashBtn2.clicked.connect(self.dashBtn2Click)
        self.dashBtn3.clicked.connect(self.dashBtn3Click)
        self.dashReconnectBtn.clicked.connect(self.dashReconnectBtnClick)
        self.CS_createBtn.clicked.connect(self.CS_createBtnClick)
        self.CS_emptyWeightInput.textChanged.connect(self.CS_weightsChanged)
        self.CS_totalWeightInput.textChanged.connect(self.CS_weightsChanged)
        self.CS_backBtn.clicked.connect(self.CS_backBtnClick)
        self.VS_backBtn.clicked.connect(self.VS_backBtnClick)
        self.dashBrowseBtn1.clicked.connect(self.dashBrowseBtn1Click)
        self.dashBrowseBtn1_2.clicked.connect(self.dashBrowseBtn2Click)
        self.VS_searchBtn.clicked.connect(self.VS_searchBtnClick)
        #self.VS_verifyShipmentBtn.clicked.connect(self.VS_verifyShipmentBtnClick)
        #self.VS_RFIDDetectBtn.clicked.connect(self.VS_RFIDDetectBtnClick)
        #self.VS_weighBtn.clicked.connect(self.VS_weighBtnClick)
        self.invoiceTable.setColumnWidth(0,110)
        self.invoiceTable.setColumnWidth(1,190)
        #self.invoiceTable.setColumnWidth(2,30)
        self.invoiceTable.setColumnWidth(2,80)
        self.invoiceTable.setColumnWidth(3,100)
        self.invoiceTable.setColumnWidth(4,100)
        self.invoiceTable.setColumnWidth(5,100)
        self.invoiceTable.setColumnWidth(6,120)
        self.invoiceTable.horizontalHeader().setVisible(True)
        self.invoiceTable.horizontalHeaderVisible = True
        self.gTotalWeight = 0
        self.dashErrorLabel.setText("")
        self.invoiceTable.itemChanged.connect(self.VS_cellChanged)
        self.dashWeightPortList.activated[str].connect(self.weighPortChanged)
        self.dashBrowseBtn3.clicked.connect(self.dashBrowseBtn3Click)
        self.fontTableHeader = QtGui.QFont()
        self.fontTableHeader.setBold(True)
        self.VS_clearBtn.clicked.connect(self.VS_clearBtnClicked)
        self.VS_weighBtn.clicked.connect(self.VS_weighBtnClicked)
        self.VS_RFIDDetectBtn.clicked.connect(self.VS_RFIDDetectBtnClicked)
        self.VS_verifyShipmentBtn.clicked.connect(self.VS_verifyShipmentBtnClicked)
        self.verificationProcess = False
        self.VS_printLabelBtn.clicked.connect(self.VS_printLabelBtnClick)
        self.dashAboutBtn.clicked.connect(self.dashAboutBtnClick)
        self.about_backbtn.clicked.connect(self.about_backbtnClick)
        self.HS_listView.itemClicked.connect(self.HS_listViewClicked)
        self.HS_backBtn.clicked.connect(self.HS_backBtnClick)
        self.HS_tableView.setColumnWidth(0,110)
        self.HS_tableView.setColumnWidth(1,190)
        self.HS_tableView.setColumnWidth(2,80)
        self.HS_tableView.setColumnWidth(3,100)
        self.HS_tableView.setColumnWidth(4,100)
        self.HS_tableView.setColumnWidth(5,100)
        self.HS_tableView.setColumnWidth(6,120)
        self.HS_tableView.horizontalHeaderVisible = True
        self.fname1 = ["",""]
        self.fname2 = ["",""]
        self.fname3 = ["",""]

    def loadingDone(self):
        print("\nloadfinish")
        if(self.CloadCount<=self.TloadCount):
            self.CloadCount = self.CloadCount+1
            print(self.CloadCount)
            self.movie.start()
        else:
            self.TloadCount=0
            self.CloadCount=0
            self.movie.stop()
            self.stackedWidget.setCurrentIndex(self.NloadIndex)

    def loginBtnClick(self):
        if(self.userInput.text()=="shadab" and self.pswdInput.text()=="shadab@123"):
            print("Login Success")
            self.label_4.setText("")
            self.stackedWidget.setCurrentIndex(1)
            self.TloadCount=5
            self.NloadIndex=2
            self.movie.start()
            self.now = datetime.now()
            self.today = date.today()
            self.dashUpdateTime()
            self.stackedWidget.setCurrentIndex(2)
            self.historyInit()



        else: 
            self.label_4.setText("Invalid Credentials!")
   
    def dashUpdateTime(self):
          self.dashDateText.setText(self.today.strftime("%d, %B %Y"))
          self.dashTimeText.setText(self.now.strftime("%H:%M"))

    def dashLogoutBtnClick(self):
        self.userInput.setText("")
        self.pswdInput.setText("")
        self.stackedWidget.setCurrentIndex(0)
    
    def dashBtn1Click(self):
        self.stackedWidget.setCurrentIndex(3)
        self.initCreateShipment()

    def dashBtn2Click(self):
        self.initVerifyShipment()

    def dashBtn3Click(self):
        self.stackedWidget.setCurrentIndex(6)
        self.historyListInit()

    def dashReconnectBtnClick(self):
        
        if not(RFID.isArduinoConnected()):
                RFID.arduinoClose()
                RFID.arduinoState=0
                RFID.arduinoConnect()
                time.sleep(1)
                if RFID.isArduinoConnected():
                        self.updateSensorStat(1,1)
                else:
                        self.updateSensorStat(1,0)
        else:
                if Weigh.weighConnect(self.weighPort):
                    self.updateSensorStat(2,1)
                else:
                    self.updateSensorStat(2,0)
        
    def updateSensorStat(self,sensor,value):
        if sensor == 1 and value == 1:
                self.rfidStatLabel.setText("CONNECTED")
                self.rfidStatLabel.setStyleSheet("color: rgb(0, 255, 0);")
        elif sensor == 1 and value ==0:
                self.rfidStatLabel.setText("NOT CONNECTED")
                self.rfidStatLabel.setStyleSheet("color: rgb(255, 0, 0);")
        else:
            if value == 1:
                self.weighScaleStatLabel.setText("CONNECTED")
                self.weighScaleStatLabel.setStyleSheet("color: rgb(0, 255, 0);")
            else:
                self.weighScaleStatLabel.setText("NOT CONNECTED")
                self.weighScaleStatLabel.setStyleSheet("color: rgb(255, 0, 0);")
    
    def initCreateShipment(self):
         Tdate = self.today.strftime("%d, %B %Y")
         print(Tdate)
         self.now = datetime.now() 
         Ttime = self.now.strftime("%H:%M")
         print(Ttime)
         self.CS_dateInput.insert(Tdate)
         self.CS_timeInput.insert(Ttime)
    
    def CS_weightsChanged(self):
        if not(self.CS_totalWeightInput.text()=="") and not(self.CS_emptyWeightInput.text()):
            self.net = float(self.CS_totalWeightInput.text())+ float(self.CS_emptyWeightInput.text())
            self.CS_netWeightLabel.setText(self.net)
    
    def CS_createBtnClick(self):
         print(self.CS_binNoInput.text())
         if (not(self.CS_binNoInput.text())=="" and not(self.CS_dateInput.text()=="") and not(self.CS_timeInput.text()=="") and not(self.CS_emptyWeightInput.text()=="") and not(self.CS_totalWeightInput.text()=="")):
              if self.CS_totalWeightInput.text().isnumeric() and self.CS_emptyWeightInput.text().isnumeric():
                   self.net = float(self.CS_totalWeightInput.text())+ float(self.CS_emptyWeightInput.text())
                   self.RFIDTagValue = self.CS_binNoInput.text() + "&&" + self.today.strftime("%d/%m/%Y")+ "&&"+ self.now.strftime("%H:%M:%S") + "&&" + self.CS_emptyWeightInput.text() + "&&"+  self.CS_totalWeightInput.text()
                   print(self.RFIDTagValue)
    
                   if RFID.isRFIDTagged():
                        self.CS_currentTag = RFID.readRFID()
                        self.CS_statusLabel.setText("RFID Tag: "+self.CS_currentTag+" Detected!.Dont Remove.")
                        self.CS_statusLabel.setStyleSheet("color: rgb(0, 0, 0);")
                        time.sleep(3)
                        if RFID.RFID_writeData(self.RFIDTagValue):
                             self.CS_statusLabel.setText("RFID WRITE SUCCESSFUL!")
                             self.CS_statusLabel.setStyleSheet("color: rgb(0, 255, 0);")
                        else:
                             self.CS_statusLabel.setText("RFID WRITE UNSUCCESSFUL")
                             self.CS_statusLabel.setStyleSheet("color: rgb(255, 0, 0);")
                   else:
                        self.CS_statusLabel.setText("Place an RFID Tag to Write.....")
                        self.CS_statusLabel.setStyleSheet("color: rgb(255, 0, 0);")
                        while not(RFID.isRFIDTagged()):
                             pass
                        self.CS_currentTag = RFID.readRFID()
                        self.CS_statusLabel.setText("RFID Tag: " + self.CS_currentTag + " Detected!. Dont Remove.")
                        self.CS_statusLabel.setStyleSheet("color: rgb(0, 0, 0);")
                        time.sleep(3)
                        if RFID.RFID_writeData(self.RFIDTagValue):
                             self.CS_statusLabel.setText("RFID WRITE SUCCESSFUL!")
                             self.CS_statusLabel.setStyleSheet("color: rgb(0, 255, 0);")
                        else:
                             self.CS_statusLabel.setText("RFID WRITE UNSUCCESSFUL")
                             self.CS_statusLabel.setStyleSheet("color: rgb(255, 0, 0);")
                        
                             

                        

              else:
                   self.CS_statusLabel.setText("Invalid Weights!")
                   self.CS_statusLabel.setStyleSheet("color: rgb(255, 0, 0);")
         else: 
                self.CS_statusLabel.setText("Missing Data!!")
                self.CS_statusLabel.setStyleSheet("color: rgb(255, 0, 0);")
    
    def CS_backBtnClick(self):
         self.stackedWidget.setCurrentIndex(2)
         self.CS_binNoInput.setText("")
         self.CS_emptyWeightInput.setText("")
         self.CS_totalWeightInput.setText("")
         self.CS_statusLabel.setText("PLEASE PLACE RFID TAG!")
         self.CS_statusLabel.setStyleSheet("color: rgb(0, 0, 0);")
 
    def VS_backBtnClick(self):
         self.stackedWidget.setCurrentIndex(2)

    def dashBrowseBtn1Click(self):
         self.fname1 = QtWidgets.QFileDialog.getOpenFileName(MainWindow, 'Open File', 'c:\\', 'Excel File (*.xlsx *.xls)')
         self.dashFileLabel1.setText(path.basename(self.fname1[0]).split('/')[-1])
        
    def dashBrowseBtn2Click(self):
         self.fname2 = QtWidgets.QFileDialog.getOpenFileName(MainWindow, 'Open File', 'c:\\', 'Excel File (*.xlsx *.xls)')
         
         self.dashTimeText_3.setText(path.basename(self.fname2[0]).split('/')[-1])
    
    def dashBrowseBtn3Click(self):
         self.fname3 = QtWidgets.QFileDialog.getOpenFileName(MainWindow, 'Open File', 'c:\\', 'Excel File (*.xlsx *.xls)')
         
         self.dashFileLabel3.setText(path.basename(self.fname3[0]).split('/')[-1])

    def initVerifyShipment(self):
         if not(self.dashFileLabel1.text()=="Choose File..") and not(self.dashTimeText_3.text()=="Choose File.." and not(self.dashFileLabel3.text()=="Choose File..")):
            if Weigh.weighState==1:
                if RFID.arduinoState==1:                         
                    self.VS_invoiceInput.setText("")
                    self.VS_resultTab.setCurrentIndex(0)
                    self.VS_weighInput.setText("")
                    self.VS_RFIDInput.setText("")
                    self.dashErrorLabel.setText("")
                    self.stackedWidget.setCurrentIndex(4)
                    self.VS_currentSearch = ""
                    self.shipmentVerified = False
                else:
                    self.RFIDDialog()
                      
            else:
                self.Dialog1 = QtWidgets.QMessageBox(MainWindow)
                self.Dialog1.setText("Are You Sure To Continue Without WeighScale? ")
                self.Dialog1.setWindowTitle("Weigh Scale Not Detected!")
                self.Dialog1.setIcon(QtWidgets.QMessageBox.Warning)
                self.Dialog1.setStandardButtons(QtWidgets.QMessageBox.Ok|QtWidgets.QMessageBox.Cancel)
                self.Dialog1.buttonClicked.connect(self.weighScaleDialog)
                self.Dialog1.exec_()


         else:
              self.dashErrorLabel.setText("Cannot Continue Without Excel Files!!")             

    def VS_searchBtnClick(self):
         
         self.masterBook = openpyxl.load_workbook(self.fname1[0])
         self.invoiceBook = openpyxl.load_workbook(self.fname2[0])
         self.binBook = openpyxl.load_workbook(self.fname3[0])

         self.invoiceSheet = self.invoiceBook.active
         self.masterSheet = self.masterBook.active
         self.binSheet = self.binBook.active

         searchText = self.VS_invoiceInput.text()

         if self.VS_currentSearch != self.VS_invoiceInput.text():
            self.searchState = True
            self.VS_searchErrorLabel.setText("")
            if searchText.__contains__(","):
                    self.columnList = searchText.split(",")
                    self.searchData = []
                    print(self.columnList)
                    self.offsetLength = {}
                    o = 0
                    for invoice in self.columnList:
                        self.tableOffset = o
                        if self.searchInvoice("InvoiceNumber", invoice.strip()):
                            self.searchData.append(self.results)
                            self.offsetLength[o] = len(self.results)
                            #print(self.searchData)
                            self.displayTable(2,0)
                            o=o+1
                        else:
                            self.VS_searchErrorLabel.setText(self.VS_searchErrorLabel.text()+" Invoice: "+ invoice + " Not Found!!")
                    #print(self.offsetLength)

            else:
                    self.columnList = searchText.split(",")
                    self.tableOffset = 0
                    if self.searchInvoice("InvoiceNumber", self.columnList[0]):
                        self.VS_searchErrorLabel.setText("")
                        self.displayTable(1,0)
                    else: 
                        self.VS_searchErrorLabel.setText("No Invoice Found!!")

                    #print(self.results)

            self.searchState = False
            self.VS_currentSearch = searchText

    def searchInvoice(self, column, tag):
            self.results = []
            for column_cell in self.invoiceSheet.iter_cols(1, self.invoiceSheet.max_column):  # iterate column cell
                        if column_cell[0].value == column:    # check for your column
                                j = 0
                                for data in column_cell[1:]:
                                        if str(data.value) == tag:
                                                partName = self.invoiceSheet.cell(data.row, column=4).value
                                                UQC = self.invoiceSheet.cell(data.row, column=6).value
                                                partNumber = self.invoiceSheet.cell(data.row, column=8).value
                                                quantity = self.invoiceSheet.cell(data.row, column=7).value
                                                balQuantity = self.invoiceSheet.cell(data.row, column=10).value
                                                if balQuantity == None:
                                                      balQuantity = quantity
                                                disQuantity = self.invoiceSheet.cell(data.row, column=9).value
                                                if disQuantity == None:
                                                      disQuantity = quantity
                                                else:
                                                      disQuantity = balQuantity
                                                net = self.searchPartWeight(str(partNumber))
                                                total = float(net) * float(balQuantity)
                                                total=round(total,4)
                                                self.results.append({"cell":data ,"offset": self.tableOffset,"slNo":j,"PartNo":partNumber, "partName":partName, "UQC":UQC, "net":net, "totalQuantity":quantity,"dispQuantity":disQuantity,"balQuantity": balQuantity,"total":total})
                                                print("Success")
                                                j=j+1
                                                self.invoiceDate = str(self.invoiceSheet.cell(data.row, column=3).value)[:10]
                                if len(self.results) == 0:
                                      return False
                                return True
    
    def displayTable(self, type, loc):
        if loc == 0:
              TABLE = self.invoiceTable
        else:
              TABLE = self.HS_tableView

        self.tableHeaderRows = []

        if type == 2:
                self.tableType = 2
                rowTotal = 0
                for i in self.searchData:
                    rowTotal = rowTotal + len(i)
                TABLE.setRowCount(rowTotal+(2*len(self.searchData))+1)
                l = 0
                i = 0
                self.binWeigh = 0
                for invoice in self.searchData:     
                    TABLE.setItem(i,0,QtWidgets.QTableWidgetItem("InvoiceNo: "))
                    TABLE.item(i,0).setFont(self.fontTableHeader)
                    TABLE.setItem(i,1,QtWidgets.QTableWidgetItem(self.columnList[l]))
                    TABLE.item(i,1).setFont(self.fontTableHeader)
                    TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(""))
                    self.tableHeaderRows.append(i)
                    l=l+1
                    i=i+1
                    self.gTotalWeight = 0
                    for row in invoice:
                                                TABLE.setItem(i,0,QtWidgets.QTableWidgetItem(str(row["PartNo"])))
                                                TABLE.setItem(i,1,QtWidgets.QTableWidgetItem(str(row["partName"])))
                                                #TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(str(row["UQC"])))
                                                TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(str(row["net"])))
                                                TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(str(row["totalQuantity"])))

                                                TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(str(row["dispQuantity"])))
                                                if row["balQuantity"] == row["totalQuantity"]:
                                                      if loc == 1:
                                                        TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                                                      else:
                                                        TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(str(row["balQuantity"])))
                                                else:
                                                      TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(str(row["balQuantity"])))
                                                
                                                TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(row["total"])))
                                                self.gTotalWeight = self.gTotalWeight + row["total"]
                                                i=i+1
                    TABLE.setItem(i,0,QtWidgets.QTableWidgetItem("TOTAL"))
                    TABLE.item(i,0).setFont(self.fontTableHeader)
                    TABLE.setItem(i,1,QtWidgets.QTableWidgetItem("WEIGHT: "))
                    TABLE.item(i,1).setFont(self.fontTableHeader)
                    TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(self.gTotalWeight)))
                    TABLE.item(i,6).setFont(self.fontTableHeader)
                    self.VS_dateInput.setText(self.invoiceDate)
                    self.tableHeaderRows.append(i)
                    self.binWeigh = self.binWeigh + self.gTotalWeight
                    i=i+1 

                TABLE.setItem(i,0,QtWidgets.QTableWidgetItem("BIN"))
                TABLE.item(i,0).setFont(self.fontTableHeader)
                TABLE.setItem(i,1,QtWidgets.QTableWidgetItem("WEIGHT: "))
                TABLE.item(i,1).setFont(self.fontTableHeader)
                TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(""))
                TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(""))
                TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(""))
                TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(self.binWeigh)))
                TABLE.item(i,6).setFont(self.fontTableHeader)

        else:       
                    self.tableType = 1
                    self.gTotalWeight = 0
                    i=0
                    TABLE.setRowCount(len(self.results)+2)
                    for row in self.results:
                                                TABLE.setItem(i,0,QtWidgets.QTableWidgetItem(str(row["PartNo"])))
                                                TABLE.setItem(i,1,QtWidgets.QTableWidgetItem(str(row["partName"])))
                                                #TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(str(row["UQC"])))
                                                TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(str(row["net"])))
                                                TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(str(row["totalQuantity"])))                                             
                                                TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(str(row["dispQuantity"])))
                                                if row["balQuantity"] == row["totalQuantity"]:
                                                      if loc == 1:
                                                        TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                                                      else:
                                                         TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(str(row["balQuantity"])))   
                                                else:
                                                      TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(str(row["balQuantity"])))
                                                TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(row["total"])))
                                                self.gTotalWeight = self.gTotalWeight + row["total"]
                                                self.gTotalWeight = round(self.gTotalWeight, 4)
                                                i=i+1
                    TABLE.setItem(i,0,QtWidgets.QTableWidgetItem("TOTAL"))
                    TABLE.item(i,0).setFont(self.fontTableHeader)
                    TABLE.setItem(i,1,QtWidgets.QTableWidgetItem("WEIGHT: "))
                    TABLE.item(i,1).setFont(self.fontTableHeader)
                    TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(self.gTotalWeight)))
                    TABLE.item(i,6).setFont(self.fontTableHeader)
                    self.VS_dateInput.setText(self.invoiceDate)
                    self.tableHeaderRows.append(i)
                    self.binWeigh = self.gTotalWeight
                    i=i+1
                    TABLE.setItem(i,0,QtWidgets.QTableWidgetItem("BIN"))
                    TABLE.item(i,0).setFont(self.fontTableHeader)
                    TABLE.setItem(i,1,QtWidgets.QTableWidgetItem("WEIGHT: "))
                    TABLE.item(i,1).setFont(self.fontTableHeader)
                    TABLE.setItem(i,2,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,3,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,4,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,5,QtWidgets.QTableWidgetItem(""))
                    TABLE.setItem(i,6,QtWidgets.QTableWidgetItem(str(self.binWeigh)))
                    TABLE.item(i,6).setFont(self.fontTableHeader)
                              
    def searchPartWeight(self,part):
                column_name = "Material"
                for column_cell in self.masterSheet.iter_cols(1, self.masterSheet.max_column):  # iterate column cell
                        if column_cell[0].value == column_name:    # check for your column
                                j = 0
                                for data in column_cell[1:]:    # iterate your column
                                        if str(data.value) == part:
                                                weight = self.masterSheet.cell(row=data.row, column=3)
                                                #print(weight.value)
                                                return weight.value
                                                
                                return 0
                        
    def searchBinNo(self, rfid):
        column_name = "RFID"
        for column_cell in self.binSheet.iter_cols(1, self.masterSheet.max_column):
 # iterate column cell
                        if column_cell[0].value == column_name:    # check for your column
                                j = 0
                                for data in column_cell[1:]:
                                        if data.value == rfid:
                                                print("bin MATCH!!")
                                                binData = str(self.binSheet.cell(row=data.row, column=2).value)+"%"+str(self.binSheet.cell(row=data.row, column=3).value)
                                                return binData
                                                
                        return "NA"
                                 
    def VS_cellChanged(self,item):
        if self.searchState == False:
            #print(item)
            if item.row() in self.tableHeaderRows:
                  pass     
            else:
                if not(self.shipmentVerified):
                    if self.tableType == 1:
                        self.searchState = True
                        for row in self.results:
                          #print(item.column())
                            if int(row["slNo"]) == item.row():
                                if item.column() == 4:
                                
                                    if int(item.text()) <= row["balQuantity"]:
                                        row["dispQuantity"] = int(item.text())
                                        row["balQuantity"] = row["balQuantity"] - row["dispQuantity"]
                                        row["total"] = row["net"] * row["dispQuantity"]
                                        row["total"] = round(row["total"], 4)
                                        excelCell = row["cell"]
                                        self.invoiceSheet.cell(row=excelCell.row, column=9).value = row["dispQuantity"]
                                        self.invoiceSheet.cell(row=excelCell.row, column=10).value = row["balQuantity"]
                                        print("change Success")
                        self.displayTable(1,0)
                        self.searchState = False
                    else: 
                        self.searchState = True
                        offset = 0
                        sl = 0
                        prev = 0
                        searching = True
                        while searching:
                    
                            curlen = self.offsetLength[offset]+2

                            if (curlen + prev) > item.row():
                                sl = item.row() - prev-1
                                searching = False
                            else:
                                prev = prev+ curlen
                                offset = offset+1 
                        #print(self.searchData[offset][sl])
                        if item.column() == 4:
                                
                                if int(item.text()) <= self.searchData[offset][sl]["totalQuantity"]:
                                    self.searchData[offset][sl]["dispQuantity"] = int(item.text())
                                    self.searchData[offset][sl]["balQuantity"] = self.searchData[offset][sl]["totalQuantity"] - self.searchData[offset][sl]["dispQuantity"]
                                    self.searchData[offset][sl]["total"] = self.searchData[offset][sl]["net"] * self.searchData[offset][sl]["dispQuantity"]
                                    self.searchData[offset][sl]["total"] = round(self.searchData[offset][sl]["total"], 4)
                                    excelCell = self.searchData[offset][sl]["cell"]
                                    self.invoiceSheet.cell(row=excelCell.row, column=9).value = self.searchData[offset][sl]["dispQuantity"]
                                    self.invoiceSheet.cell(row=excelCell.row, column=10).value = self.searchData[offset][sl]["balQuantity"]
                                    print("change Success")
                                    print(self.searchData[offset][sl])
                                    self.displayTable(2,0)
                        self.searchState = False
                else:
                      self.VS_ErrorLabel_2.setText("CANNOT CHANGE DATA AFTER VERIFICATION! START NEW!!")
                      self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 255, 255);")

    def weighPortChanged(self, port):
          print(port)
          self.weighPort = port
         
    def VS_verifyShipmentBtnClicked(self):
        #if RFID.isRFIDTagged():pass
        #self.JSON_save()

        if RFID.isArduinoConnected():
            self.VS_progress.setValue(20)
            self.VS_resultTab.setCurrentIndex(2)
            if RFID.isRFIDTagged():
                self.VS_ErrorLabel_2.setText("")
                self.VS_progress.setValue(30)
                self.curRFID = RFID.readRFID()
                self.VS_RFIDInput.setText(self.curRFID)
                self.VS_progress.setValue(40)
                if self.curRFID != "NA":
                    self.VS_progress.setValue(50)
                    data=self.searchBinNo(self.curRFID)
                    self.VS_progress.setValue(60)
                    if data.__contains__("%"):
                        binData=data.split("%")
                        self.curBinNo = binData[0]
                        self.curBinEmptyWeigh = float(binData[1])
                        self.VS_binNoInput.setText(self.curBinNo)
                        self.VS_emptyBinWeighInput.setText(self.curBinEmptyWeigh)
                        print(self.curBinEmptyWeigh)
                    else:
                          self.VS_ErrorLabel_2.setText("No Bin Data Found For Tag: "+ self.curRFID)
                          self.VS_progress.setValue(0)
                          return
                    self.VS_progress.setValue(70)                  
                    self.VS_ErrorLabel_2.setText("Please Enter Weigh Scale Value!!")
                    self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 255, 255);")
                    self.VS_weighInput.setReadOnly(False)
                    self.VS_weighInput.setText("")
                    self.verificationProcess = True

                            
            else:
                self.VS_ErrorLabel_2.setText("RFID Not Tagged!! PLEASE TAG.")
                self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 0, 0);")
                self.VS_progress.setValue(0)
                
        else:
            self.VS_ErrorLabel_2.setText("RFID not Connected!! PLEASE RETRY!!")
            self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 0, 0);") 
            self.VS_progress.setValue(0)            

          #self.labelPDFInit()
          #self.generatePDF()

    def VS_clearBtnClicked(self):
          self.invoiceTable.setRowCount(0)
          self.VS_resultTab.setCurrentIndex(0)
          self.VS_invoiceInput.setText("")
          self.VS_dateInput.setText("")
          self.VS_currentSearch =  ""
          self.VS_weighInput.setText("")
          self.VS_RFIDInput.setText("")
          self.VS_binNoInput.setText("")
          self.VS_emptyBinWeighInput.setText("")
          self.shipmentVerified =False
          self.VS_ErrorLabel_2.setText("")
          self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 255, 255);")

    def VS_weighBtnClicked(self):
        if self.verificationProcess:
            if (self.VS_weighInput.text()).replace('.','',1).isdigit():
                  self.VS_ErrorLabel_2.setText("")
                  self.VS_progress.setValue(80) 
                  netWeigh = float(self.VS_weighInput.text()) - self.curBinEmptyWeigh
                  #HEREEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEEE
                  if isclose(float(netWeigh), self.binWeigh, abs_tol=3):
                             print("VERIFICATION SUCCESS!!")
                             self.seqCount=self.seqCount+1
                             self.curSeqNo = self.seqCount
                             self.VS_progress.setValue(100) 
                             self.VS_ErrorLabel_2.setText("VERIFICATION SUCCESSFUL!!")
                             self.VS_ErrorLabel_2.setStyleSheet("color: rgb(0, 255, 0);")
                             self.shipmentVerified = True
                             self.verificationSuccessDialog(str(round(netWeigh,2))+" ≈≈ "+str(self.binWeigh))
                             self.invoiceBook.save(self.fname2[0])
                             
                            #JSON STUFFF
                             self.JSON_save()

                             self.VS_resultTab.setCurrentIndex(1)

                  else:
                        self.shipmentVerified = False
                        self.VS_progress.setValue(0)
                        print("VERIFICATION FAILED!!")
                        self.VS_ErrorLabel_2.setText("VERIFICATION UNSUCCESSFUL!!")
                        self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 0, 0);")
                        self.shipmentNotVerified()

            else:
                  self.VS_ErrorLabel_2.setText("INVALID WEIGH INPUT!!")
                  self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 0, 0);")
                  self.VS_progress.setValue(0)

        else:  
            if Weigh.weighState:
                self.VS_weighInput.setText(Weigh.weighRead())
                self.VS_ErrorLabel_2.setText("")
            else:
                self.VS_ErrorLabel_2.setText("WEIGH SCALE NOT CONNECTED!!")
                self.VS_ErrorLabel_2.setStyleSheet("color: rgb(255, 255, 255);")
    
    def VS_RFIDDetectBtnClicked(self):
          if RFID.isArduinoConnected():
                if RFID.isRFIDTagged():
                      self.VS_RFIDInput.setText(RFID.readRFID())
                      self.VS_ErrorLabel_2.setText("")

                else:
                      self.VS_ErrorLabel_2.setText("PLEASE PLACE RFID TAG!!")
          else:
                self.VS_ErrorLabel_2.setText("RFID NOT CONNECTED!!")

    def VS_verify(self):
        pass      

    def labelPDFInit(self):
        self.env = Environment(loader=FileSystemLoader('.'))
        self.label = self.env.get_template("label.html")

        self.html_string = self.label.render(self.labelInput)

        #self.path_wkhtmltopdf = path_wkhtmltopdf_input
        self.config = pdfkit.configuration(wkhtmltopdf=r"C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe")

        self.options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'enable-local-file-access': None,
            'no-outline': None,
            'orientation':'landscape'
            }
    
    def generatePDF(self):
          pdfkit.from_string(self.html_string
                   , ('labelOutputs/'+str(self.curBinNo)+'.pdf')
                   , configuration = self.config
                   , options = self.options
                   )

    def VS_printLabelBtnClick(self):
        invoices = ""
        for inv in self.columnList:
                invoices = invoices + ", "+ str(inv)
        self.labelInput = {
                "binNo" : self.curBinNo,
                "seqNo" : self.curSeqNo,
                "date" : self.today.strftime("%d, %B %Y"),
                "time" : self.now.strftime("%H:%M"),
                "invoiceNo" : invoices[1:],
                "binWeigh" : (str(round((self.binWeigh + self.curBinEmptyWeigh),2))+" Kgs")
          }
        self.labelPDFInit()
        self.generatePDF()
        print("printing PDF: ")
        ROOT_DIR = path.dirname(path.abspath(__file__))
        loc = "labelOutputs\\"  + str(self.curBinNo) + ".pdf"
        system(loc)  
        time.sleep(2)
        self.VS_clearBtnClicked()
   
    def RFIDDialog(self):
            self.Dialog2 = QtWidgets.QMessageBox(MainWindow)
            self.Dialog2.setText("Please Connect The RFID Device to Continue!!")
            self.Dialog2.setWindowTitle("RFID Not Connected!!")
            self.Dialog2.setIcon(QtWidgets.QMessageBox.Critical)
            self.Dialog2.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.Dialog2.exec_()

    def weighScaleDialog(self,i):
          if i.text() == "OK":
                if RFID.arduinoState==1:                         
                    self.VS_invoiceInput.setText("")
                    self.VS_resultTab.setCurrentIndex(0)
                    self.VS_weighInput.setText("")
                    self.VS_RFIDInput.setText("")
                    self.VS_binNoInput.setText("")
                    self.VS_emptyBinWeighInput.setText("")
                    self.dashErrorLabel.setText("")
                    self.stackedWidget.setCurrentIndex(4)
                    self.VS_currentSearch = ""
                    self.shipmentVerified = False
                else:
                    self.RFIDDialog()

    def verificationSuccessDialog(self, Value):
            self.Dialog3 = QtWidgets.QMessageBox(MainWindow)
            self.successLogo = QtGui.QPixmap('Resouces/success.jpg')
            self.Dialog3.setIconPixmap(self.successLogo)
            self.Dialog3.setWindowTitle("Verification Success!!")
            self.Dialog3.setText("SUCCESSFUL!! Bin Match With Invoice to: "+ Value)
          
            self.Dialog3.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.Dialog3.exec_()

    def shipmentNotVerified(self):
            self.Dialog3 = QtWidgets.QMessageBox(MainWindow)
            #self.Dialog3.setIconPixmap(self.successLogo)
            self.Dialog3.setWindowTitle("Verification Failed!!")
            self.Dialog3.setText("Bin Weigh does not Match with Invoice!! Rework!!!!")
            self.VS_resultTab.setCurrentIndex(0)
            self.Dialog3.setStandardButtons(QtWidgets.QMessageBox.Ok)
            self.Dialog3.exec_()
          
    def about_backbtnClick(self):
          self.stackedWidget.setCurrentIndex(2)
    
    def dashAboutBtnClick(self):
          self.stackedWidget.setCurrentIndex(5)

    def historyInit(self):
        self.h = open('appDat.json')

        self.jsonData = json.load(self.h)

        if self.jsonData == '{}':
                return

       
                #print(i['name'])
          
        if  self.jsonData['FileData']['f1'] !="":
                        self.fname1[0] = self.jsonData['FileData']['f1']
                        self.dashFileLabel1.setText(path.basename(self.fname1[0]).split('/')[-1])
        
                        self.fname2[0] = self.jsonData['FileData']['f2']
                        self.dashTimeText_3.setText(path.basename(self.fname2[0]).split('/')[-1])

                        self.fname3[0] = self.jsonData['FileData']['f3']
                        self.dashFileLabel3.setText(path.basename(self.fname3[0]).split('/')[-1])

        self.seqCount = self.jsonData['SequenceCounter']
        
        self.h.close()
    
    def historyListInit(self):
        self.HS_listView.clear()
        self.HS_binNoInput.setText("")
        self.HS_emptyWeighInput.setText("")
        self.HS_seqNoInput.setText("")

        for i in self.jsonData['History']:
                self.HS_listView.addItem(i['name']) 

    def JSON_save(self):
            if len(self.columnList)>1:
                                    invoices = ""
                                    for inv in self.columnList:
                                        invoices = invoices + ", "+ str(inv)
                                        searchData = self.searchData
                                        for i in searchData:
                                              for j in i:
                                                    #print(j['cell'])
                                                    j.pop('cell','NO CELL FPPUND!')
                                        historyData = {"type":2, "invoice":invoices[1:], "data":searchData, "name":(invoices[1:]+"- "+self.today.strftime("%d, %B %Y")+"- "+self.now.strftime("%H:%M")), "binNo": self.curBinNo, "emptyBinWeigh": self.curBinEmptyWeigh, "seqNo":self.curSeqNo}
            else:
                                   searchDat = self.results
                                   for i in searchDat:
                                              del i['cell']
                                   historyData = {"type":1, "invoice":self.columnList[0],"data":searchDat, "name":(self.columnList[0])+"- "+self.today.strftime("%d, %B %Y")+"- "+self.now.strftime("%H:%M"), "binNo": self.curBinNo, "emptyBinWeigh": self.curBinEmptyWeigh, "seqNo":self.curSeqNo}
                        
            self.jsonData['FileData'].update({"f1":self.fname1[0]})
            self.jsonData['FileData'].update({"f2":self.fname2[0]})  
            self.jsonData['FileData'].update({"f3":self.fname3[0]}) 
            self.jsonData['sequenceCounter'].update(self.seqCount)                   
            self.jsonData['History'].append(historyData)
            with open("AppDat.json", "w") as outJson:
                                json.dump(self.jsonData, outJson, indent=4)
                                print("json Suucess")
       
    def HS_backBtnClick(self):
          self.stackedWidget.setCurrentIndex(2)

    def HS_listViewClicked(self, item):
          rowNum=self.HS_listView.row(item)
          if self.jsonData['History'][rowNum]['type']==1:
                type = 1
                self.results = self.jsonData['History'][rowNum]['data']
          else:
                type = 2
                self.searchData = self.jsonData['History'][rowNum]['data']
                self.columnList = self.jsonData['History'][rowNum]['invoice'].split(",")
          self.invoiceDate = self.jsonData['History'][rowNum]['name'].split("-")[1]
          self.displayTable(type,2)
          self.HS_binNoInput.setText(str(self.jsonData['History'][rowNum]['binNo']))
          self.HS_emptyWeighInput.setText(str(self.jsonData['History'][rowNum]['emptyBinWeigh']))
          self.HS_seqNoInput.setText(str(self.jsonData['History'][rowNum]['seqNo']))   

      
if __name__ == "__main__":
    import sys
    winProcess = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    
    App = MainApp()
    App.setupUi(MainWindow)
    App.gui_append()

    RFID = arduino()
    Weigh = weighScale()

    App.stackedWidget.setCurrentIndex(0)

    if RFID.arduinoState == 1:
                print("ardiono Connected")
                App.updateSensorStat(1,1)
    else:
                print("arduino Not Connected")
                App.updateSensorStat(1,0)
  
    for p in Weigh.portList:
          App.dashWeighPortList.addItem(str(p))
    
    MainWindow.show()
    sys.exit(winProcess.exec_())


